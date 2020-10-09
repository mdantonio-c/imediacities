#!.venv/bin/python3
import json
import sys
from datetime import datetime
from functools import reduce

from openpyxl import load_workbook

imc_namespace = "https://imediacities.eu/resource/"
replaces = (" ", "_"), ("/", "_")


def generate_iri(term):
    return imc_namespace + reduce(lambda a, kv: a.replace(*kv), replaces, term.strip())


def lookup_term_class(start_with="Category"):
    labels = dict()
    for row in ws.iter_rows(min_row=1, max_col=13, max_row=1):
        for c in [cell for cell in row if cell.value.startswith(start_with)]:
            class_label = ws.cell(row=c.row + 1, column=c.column).value
            if class_label:
                labels[c.value.rsplit("_", 1)[1].lower()] = class_label.strip()
    return labels


def lookup_iri(idx, header="URI"):
    for row in ws.iter_rows(min_row=1, max_col=13, max_row=1):
        for c in [cell for cell in row if cell.value == header]:
            term_iri = ws.cell(row=idx, column=c.column).value
            if term_iri:
                return term_iri.strip()


def lookup_labels(idx, start_with, capitalize=False):
    labels = dict()
    for row in ws.iter_rows(min_row=1, max_col=13, max_row=1):
        for c in [cell for cell in row if cell.value.startswith(start_with)]:
            label = ws.cell(row=idx, column=c.column).value
            if label:
                labels[c.value.rsplit("_", 1)[1].lower()] = (
                    str(label).strip()
                    if not capitalize
                    else str(label).strip().capitalize()
                )
    return labels


input_filename = "IMC-Vocabularies-Manual_Annotation.xlsx"
wb = None
try:
    wb = load_workbook(filename=input_filename)
except FileNotFoundError:
    sys.exit(f"ERROR: File '{input_filename}' not found")
vocabulary = {
    "terms": [],
    "vocabulary": "IMC controlled vocabulary",
    "created": str(datetime.now()),
}
for ws in wb.worksheets:
    if ws.sheet_state == "hidden":
        print(f"WARN: ignore hidden sheet: {ws.title}")
        continue

    vocabulary_class = {"label": lookup_term_class(), "children": []}
    # print(json.dumps(vocabulary_class, indent=2, sort_keys=True))

    group_terms = {}
    idx_count = 0
    for row_index in range(2, ws.max_row + 1):
        idx_count += 1
        # get List_Entry
        entry_labels = lookup_labels(row_index, "List_Entry")
        if not entry_labels:
            # ignore groups without entry value
            continue

        # get IRI
        iri = lookup_iri(row_index)
        if iri is None:
            iri = generate_iri(entry_labels["en"])
        entry = {"id": iri, "label": entry_labels}

        # group = ws["I" + str(row_index)].value
        group_labels = lookup_labels(row_index, "Subgroup", capitalize=True)
        if group_labels:
            g = group_labels.get("en")
            if not g:
                sys.exit(
                    f"ERROR: Invalid input file. Missing 'en' term group for class '{vocabulary_class['label']['en']}'."
                    f" Current values: {group_labels}"
                )
            # print("adding entry to group '{}'".format(g))
            actual_terms = group_terms.get(g, None)
            if actual_terms is None:
                group_terms[g] = (group_labels, [entry])
            else:
                group_terms[g][1].append(entry)
        else:
            vocabulary_class["children"].append(entry)

    for key in group_terms:
        vocabulary_class["children"].append(
            {"label": group_terms[key][0], "children": group_terms[key][1]}
        )
    vocabulary["terms"].append(vocabulary_class)

# json_data = json.dumps(vocabulary, indent=2)
# print(json_data)

with open("vocabulary.json", "w") as f:
    json.dump(vocabulary, f, sort_keys=True, indent=2, ensure_ascii=False)
