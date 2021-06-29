#!.venv/bin/python3
import argparse
import copy
import datetime
import os
import sys
from distutils.util import strtobool
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

EFG_NAMESPACE = "http://www.europeanfilmgateway.eu/efg"
XSI_NAMESPACE = "http://www.w3.org/2001/XMLSchema-instance"
XSD_NAMESPACE = "http://www.w3.org/2001/XMLSchema"

ET.register_namespace("", EFG_NAMESPACE)

generated_on = str(datetime.datetime.now())
ALLOWED_ITEM_TYPES = ("IMAGE", "VIDEO", "3D-MODEL")
COMMON_MANDATORY_COLS = [
    "identifier",
    "sourceID",
    "provider",
    "provider_id",
    "rightsStatus",
]

parser = argparse.ArgumentParser()
parser.add_argument("input", help="input excel file")
parser.add_argument(
    "-t", "--target-directory", help="put all generated files into TARGET DIRECTORY"
)
parser.add_argument("-p", "--prefix", help="set a prefix to output filename")
args = parser.parse_args()
input_filename = args.input
# check target directory
target_dir = args.target_directory
if target_dir and not os.path.isdir(target_dir):
    exit(f"ERROR - Target director '{target_dir}' does NOT exist")

wb = None
try:
    wb = load_workbook(filename=input_filename, read_only=False, data_only=True)
except (InvalidFileException, FileNotFoundError) as e:
    sys.exit(f"ERROR - {e}")


def indent(elem, level=0):
    i = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


def is_blank(my_string):
    return not (my_string and str(my_string).strip())


def lookup_fields(row_idx, start_with, next_col=None):
    res = {}
    for header_row in ws.iter_rows(min_row=1, max_row=1):
        for c in [
            x
            for x in header_row
            if x.value is not None and x.value.startswith(start_with)
        ]:
            header_val = c.value.strip()
            key = header_val[len(start_with) + 1 :]
            val = ws.cell(row=row_idx, column=c.column).value
            next_val = None
            if next_col and ws.cell(row=c.row, column=c.column + 1).value == next_col:
                next_val = ws.cell(row=row_idx, column=c.column + 1).value
            res[key] = (
                str(val).strip() if val else None,
                str(next_val).strip() if next_val else None,
            )
    return res


def parse_related_entities(row_idx, el_name, parent):
    agents = ws.cell(row=row_idx, column=headers[el_name]).value
    if agents is None:
        return
    for agent in agents.split(sep=";"):
        if is_blank(agent):
            continue
        agent = agent.strip()
        rel_agent = ET.SubElement(parent, el_name)
        agent_id, agent_name, agent_type = None, None, None
        try:
            tokens = agent.strip().split(":", 2)
            agent_id = tokens[0]
            agent_name = tokens[1]
            agent_type = tokens[2]
        except IndexError:
            pass
        ET.SubElement(rel_agent, "identifier").text = agent_id.strip()
        if not agent_name:
            print(
                f"WARNING - Record[{row_idx}]. Expected name in form of 'identifier:name:?type'. Invalid value for '{agent}'"
            )
            continue
        ET.SubElement(rel_agent, "name").text = agent_name.strip()
        if agent_type:
            ET.SubElement(rel_agent, "type").text = agent_type.strip()


def add_coverage(row_idx, manifestation):
    coverage = lookup_fields(row_idx, start_with="coverage")
    coverage_el = ET.SubElement(manifestation, "coverage")
    for key, val in coverage.items():
        # expected key as follows:
        # - coverage_spatial_latlng
        # - coverage_spatial_altitude
        # - coverage_temporal
        if is_blank(val[0]):
            continue
        if key.startswith("spatial"):
            if len(key.split("_")) != 2:
                # ignore invalid coverage definitions for spatial
                continue
            spatial_type = key.split("_")[1]
            if spatial_type not in ["latlng", "altitude"]:
                continue
            spatial_el = ET.SubElement(coverage_el, "spatial")
            spatial_el.text = val[0]
            spatial_el.set("type", spatial_type)
        elif key == "temporal" and not is_blank(val[0]):
            ET.SubElement(coverage_el, "temporal").text = val[0]
    # eventually remove empty coverage element
    if coverage_el.find("spatial") is None and coverage_el.find("temporal") is None:
        manifestation.remove(coverage_el)


def add_digital_format(row_idx, manifestation):
    if headers.get("digitalFormat") and not is_blank(
        digital_format := ws.cell(row=row_idx, column=headers["digitalFormat"]).value
    ):
        digital_format_el = ET.SubElement(manifestation, "digitalFormat")
        digital_format_el.text = digital_format.strip()
        if headers.get("digitalSize"):
            digital_size = ws.cell(row=row_idx, column=headers["digitalSize"]).value
            if not isinstance(digital_size, int):
                print(
                    f"WARNING - Digital size MUST be a valid number. Actual value: '{digital_size}'"
                )
                return
            digital_format_el.set("size", str(digital_size))


def add_rights_holder(row_idx, manifestation):
    if headers.get("rightsHolder") and not is_blank(
        rights_holders := ws.cell(row=row_idx, column=headers["rightsHolder"]).value
    ):
        for rh in rights_holders.split(sep=";"):
            if is_blank(rh):
                continue
            rights_holder = ET.SubElement(manifestation, "rightsHolder")
            rights_holder.text = rh.strip()


def add_keywords(row_idx, creation):
    keywords = lookup_fields(row_idx, start_with="keywords")
    for key, val in keywords.items():
        # expected key is as follows: Subject_it, Place_it
        if len(key.split("_")) != 2:
            # ignore invalid keywords definitions
            continue
        if is_blank(val[0]):
            # no keywords for this type_lang column
            continue
        k_type, lang = key.split("_")
        keywords_el = ET.SubElement(creation, "keywords")
        keywords_el.set("lang", lang)
        keywords_el.set("type", k_type)
        for term in val[0].split(";"):
            if is_blank(term):
                continue
            ET.SubElement(keywords_el, "term").text = term.strip()


def add_descriptions(row_idx, creation):
    descriptions = lookup_fields(row_idx, start_with="description")
    for key, val in descriptions.items():
        if is_blank(val[0]):
            continue
        description_el = ET.SubElement(creation, "description")
        description_el.set("lang", key)
        description_el.text = val[0]


def create_record():
    efg_entity = ET.Element(f"{{{EFG_NAMESPACE}}}efgEntity")
    c_tag = "avcreation" if ws.title == "Video" else "nonavcreation"
    creation = ET.SubElement(efg_entity, c_tag)
    row_idx = row[0].row

    # identifier: attribute scheme= (1)
    identifier = ET.SubElement(creation, "identifier", scheme="CP_CATEGORY_ID")
    identifier.text = ws.cell(row=row_idx, column=headers["identifier"]).value

    # recordSource (1-N)
    record_source = ET.SubElement(creation, "recordSource")
    # source_id = lookup_field(row_idx, "sourceID")
    source_id = ws.cell(row=row_idx, column=headers["sourceID"]).value
    if is_blank(source_id):
        raise ValueError("missing sourceID")
    source_id = source_id.strip()
    ET.SubElement(record_source, "sourceID").text = source_id
    provider_id = ws.cell(row=row_idx, column=headers["provider_id"]).value
    if is_blank(provider_id):
        raise ValueError("missing provider_id")
    provider_id = provider_id.strip()
    provider = ws.cell(row=row_idx, column=headers["provider"]).value
    if is_blank(provider):
        raise ValueError("missing provider")
    provider = provider.strip()
    ET.SubElement(
        record_source, "provider", schemeID="Institution acronym", id=provider_id
    ).text = provider

    # title (1-N)
    titles = lookup_fields(row_idx, start_with="title_text", next_col="title_relation")
    for key, val in titles.items():
        if is_blank(val[0]):
            continue
        title_el = ET.SubElement(creation, "title")
        title_el.set("lang", key)
        ET.SubElement(title_el, "text").text = val[0]
        if val[1]:
            ET.SubElement(title_el, "relation").text = val[1]

    # identifyingTitle (1)
    identifying_title = ET.SubElement(creation, "identifyingTitle")
    identifying_title.text = (
        ws.cell(row=row_idx, column=headers["identifyingTitle"]).value
        if headers.get("identifyingTitle")
        else list(titles.values())[0][0]
    )

    if ws.title == "Video":
        # countryOfReference (1-N)
        countries = ws.cell(row=row_idx, column=headers["countryOfReference"]).value
        if is_blank(countries):
            raise ValueError("Missing countryOfReference")
        for c in countries.split(sep=";"):
            country_of_reference = ET.SubElement(creation, "countryOfReference")
            country_of_reference.text = c.strip()

        # productionYear (1-N)
        years = ws.cell(row=row_idx, column=headers["productionYear"]).value
        if is_blank(years):
            raise ValueError("Missing productionYear")
        for y in str(years).split(sep=";"):
            production_year = ET.SubElement(creation, "productionYear")
            production_year.text = y.strip()
    else:
        # dateCreated
        if "dateCreated" in headers:
            date_created = ws.cell(row=row_idx, column=headers["dateCreated"]).value
            if date_created and (
                not isinstance(date_created, str) or not is_blank(date_created)
            ):
                if isinstance(date_created, int):
                    date_created = str(date_created)
                ET.SubElement(creation, "dateCreated").text = (
                    date_created.strip()
                    if isinstance(date_created, str)
                    else date_created.strftime("%Y-%m-%d")
                )
        if "dateIssued" in headers:
            date_issued = ws.cell(row=row_idx, column=headers["dateIssued"]).value
            if date_issued and (
                not isinstance(date_issued, str) or not is_blank(date_issued)
            ):
                if isinstance(date_created, int):
                    date_issued = str(date_issued)
                date_el = ET.SubElement(creation, "date")
                date_el.text = (
                    date_issued.strip()
                    if isinstance(date_issued, str)
                    else date_issued.strftime("%Y-%m-%d")
                )
                date_el.set("type", "issued")

    # keywords (0-N)
    add_keywords(row_idx, creation)
    # description (0-N)
    add_descriptions(row_idx, creation)
    # avManifestation (1-N)
    m_tag = "avManifestation" if ws.title == "Video" else "nonAVManifestation"
    manifestation = ET.SubElement(creation, m_tag)
    # identifier
    manifestation.append(copy.deepcopy(identifier))
    # recordSource
    manifestation.append(copy.deepcopy(record_source))
    # title
    title_node = creation.find("./title[1]")
    if title_node:
        manifestation.append(copy.deepcopy(title_node))
    # language (0-N)
    if "language" in headers:
        languages = ws.cell(row=row_idx, column=headers["language"]).value
        if languages:
            for lang in languages.split(sep=";"):
                if is_blank(lang):
                    continue
                language = ET.SubElement(manifestation, "language")
                language.text = lang.strip()
    if m_tag == "avManifestation":
        # duration (0-1)
        if "duration" in headers:
            duration = ws.cell(row=row_idx, column=headers["duration"]).value
            if duration:
                ET.SubElement(manifestation, "duration").text = f"{duration}"
        # format (0-1)
        av_format = ET.SubElement(manifestation, "format")
        # gauge (0-1)
        if "gauge" in headers:
            gauge = ws.cell(row=row_idx, column=headers["gauge"]).value
            if not is_blank(gauge):
                ET.SubElement(av_format, "gauge").text = gauge.strip()
        # aspectRatio (0-1)
        if "aspectRatio" in headers:
            aspect_ratio = ws.cell(row=row_idx, column=headers["aspectRatio"]).value
            if not is_blank(aspect_ratio):
                ET.SubElement(av_format, "aspectRatio").text = aspect_ratio.strip()
        # colour (0-1)
        if "colour" in headers:
            colour = ws.cell(row=row_idx, column=headers["colour"]).value
            if not is_blank(colour):
                ET.SubElement(av_format, "colour").text = colour.strip()
        # sound (0-1)
        if "hasSound" in headers:
            has_sound = ws.cell(row=row_idx, column=headers["hasSound"]).value
            if has_sound is not None and not is_blank(f"{has_sound}"):
                has_sound = strtobool(f"{has_sound}")
                sound_el = ET.SubElement(av_format, "sound")
                sound_el.text = "With sound" if bool(has_sound) else "Without sound"
                sound_el.set("hasSound", "true" if bool(has_sound) else "false")
    if m_tag == "nonAVManifestation":
        non_av_type = ws.title.lower()
        ET.SubElement(manifestation, "type").text = non_av_type
        if non_av_type != "3d-model":
            specific_type = ws.cell(row=row_idx, column=headers["specificType"]).value
            if is_blank(specific_type):
                raise ValueError(f"Missing SpecificType for type {non_av_type}")
            ET.SubElement(manifestation, "specificType").text = specific_type.strip()
        # digitalFormat (0-1)
        add_digital_format(row_idx, manifestation)
        if "physicalFormat" in headers and not is_blank(
            physical_format := ws.cell(
                row=row_idx, column=headers["physicalFormat"]
            ).value
        ):
            ET.SubElement(
                manifestation, "physicalFormat"
            ).text = physical_format.strip()
        if "colour" in headers and not is_blank(
            colour := ws.cell(row=row_idx, column=headers["colour"]).value
        ):
            ET.SubElement(manifestation, "colour").text = colour.strip()
    # coverage (0-1)
    add_coverage(row_idx, manifestation)
    # rightsHolder (0-N)
    add_rights_holder(row_idx, manifestation)
    # rightsStatus (1)
    rights_status = ws.cell(row=row_idx, column=headers["rightsStatus"]).value
    if is_blank(rights_status):
        raise ValueError("Missing rightsStatus")
    ET.SubElement(manifestation, "rightsStatus").text = rights_status.strip()
    # thumbnail (0-1)
    if headers.get("thumbnail"):
        thumbnail_url = ws.cell(row=row_idx, column=headers["thumbnail"]).value
        if not is_blank(thumbnail_url):
            ET.SubElement(manifestation, "thumbnail").text = thumbnail_url.strip()
    if m_tag == "avManifestation" and "isShownAt" in headers:
        # item/isShownAt (0-1)
        is_shown_at = ws.cell(row=row_idx, column=headers["isShownAt"]).value
        if not is_blank(is_shown_at):
            item = ET.SubElement(manifestation, "item")
            ET.SubElement(item, "isShownAt").text = is_shown_at.strip()
        # item/type (0-1)
        ET.SubElement(manifestation, "type").text = "video"
    # relPerson (0-N)
    if headers.get("relPerson"):
        parse_related_entities(row_idx, "relPerson", creation)
    # relCorporate (0-N)
    if headers.get("relCorporate"):
        parse_related_entities(row_idx, "relCorporate", creation)
    # relCollection (0-N)
    if headers.get("relCollection"):
        parse_related_entities(row_idx, "relCollection", creation)

    indent(efg_entity)

    # wrap it in an ElementTree instance, and save as XML
    tree = ET.ElementTree(efg_entity)
    global target_dir
    if target_dir is None:
        target_dir = ""
    prefix = ""
    if args.prefix:
        prefix = f"{args.prefix}_"
    tree.write(
        os.path.join(target_dir, f"{prefix}{source_id}.xml"),
        xml_declaration=True,
        encoding="utf-8",
        method="xml",
    )


def check_mandatory_columns():
    for col in COMMON_MANDATORY_COLS:
        if col not in headers:
            raise ValueError(f"Missing mandatory column <{col}>")
    if not [n for n in headers.keys() if n and n.startswith("title_text_")]:
        raise ValueError("Missing mandatory column <title_text>")
    if ws.title == "Video":
        for col in ["countryOfReference", "productionYear"]:
            if col not in headers:
                raise ValueError(f"Missing mandatory column <{col}>")
    elif ws.title == "Image":
        for col in ["specificType"]:
            if col not in headers:
                raise ValueError(f"Missing mandatory column <{col}>")


# iterate over worksheets
for ws in [a for a in wb.worksheets if a.title.strip().upper() in ALLOWED_ITEM_TYPES]:
    print(f"Worksheet {ws.title}")
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            headers[cell.value] = cell.column
    try:
        check_mandatory_columns()
    except ValueError as err:
        print(f"ERROR\t- Worksheet {ws.title} cannot be parsed. Reason: {err}")
        continue
    counter = 0
    err_counter = 0
    for row in ws.iter_rows(2, ws.max_row + 1):
        if row[0].value is not None:
            counter += 1
            try:
                create_record()
                print(f"SUCCESS\t- Record[idx={row[0].row}] created")
            except ValueError as err:
                err_counter += 1
                print(f"ERROR\t- Record[idx={row[0].row}] cannot be parsed: {err}")

    print(
        f"Total {ws.title} items: {counter} [SUCCESS:{counter-err_counter}, FAILURE:{err_counter}]"
    )
